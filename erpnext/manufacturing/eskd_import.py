"""Import the ЄСКД workbook (`ЄСКД.xlsx`) into ESKD Product / ESKD Document / Specification.

MANUAL TOOL — never wire this into a hook, a patch or the scheduler. The catalog is
maintained by hand through the Specification templates; this module exists for the one-off
bulk load of a workbook and defaults to a dry run.

Run from the container console:

	bench --site frontend execute erpnext.manufacturing.eskd_import.run \
		--kwargs "{'path': '/tmp/ЄСКД.xlsx', 'dry_run': True}"

The import is idempotent: products, documents and specifications are matched on their
natural key (product name / code+product / specification code) and updated in place, so
re-running after the workbook changes only applies the delta.
"""

import re

import frappe
from frappe.utils import cint, flt

PLACEHOLDER_RE = re.compile(r"[?ХX]{2,}")
ORG_PREFIX_RE = re.compile(r"^([А-ЯІЇЄҐA-Z]{4})[.\s]")

# "125 0,25 5 км" / "150 0,2 30 км" -> spool diameter, fibre diameter, winding length
COIL_PURPOSE_RE = re.compile(r"^(\d+)\s+([\d.,]+)\s+([\d.,]+)\s*км")
# "ДШВ 1.5 км"
COIL_SHORT_RE = re.compile(r"([\d.,]+)\s*км")

PARAM_WINDING_LENGTH = "Довжина намотування, км"
PARAM_SPOOL_DIAMETER = "Діаметр шпулі, мм"
PARAM_FIBRE_DIAMETER = "Діаметр волокна, мм"
PARAM_CAMERA_CHANNEL = "Тип каналу камери"
PARAM_CAMERA_SIGNAL = "Тип сигналу камери"
PARAM_BATTERY_LAYOUT = "Конфігурація батареї"

DOCUMENT_TYPES = [
	{"type_name": "Специфікація", "abbreviation": "С"},
	{"type_name": "Технічні умови", "abbreviation": "ТУ"},
	{"type_name": "Інструкція користувача", "abbreviation": "ІК"},
	{"type_name": "Паспорт", "abbreviation": "ПС"},
	{"type_name": "Складальний кресленик", "abbreviation": "СК"},
	{"type_name": "Схема електрична", "abbreviation": "ЭХ"},
	{"type_name": "Технологічна карта", "abbreviation": "ТК"},
	{"type_name": "Деталь", "abbreviation": ""},
]


class Summary:
	def __init__(self):
		self.rows = {}

	def hit(self, key, delta=1):
		self.rows[key] = self.rows.get(key, 0) + delta

	def as_dict(self):
		return dict(sorted(self.rows.items()))

	def show(self):
		for key, count in sorted(self.rows.items()):
			print(f"  {key}: {count}")


def _norm(value):
	if value is None:
		return ""
	return str(value).strip()


def _is_placeholder(code):
	"""`УКРП.430103.ХХХ ЭХ` / `УКРП.200121.1001??0021С` are slots, not designations."""
	return bool(PLACEHOLDER_RE.search(code.upper()))


def _org_code(code, default="УКРП"):
	match = ORG_PREFIX_RE.match(code)
	return match.group(1) if match else default


def _num(value):
	text = _norm(value).replace(",", ".")
	if not text:
		return None
	try:
		return flt(text)
	except (ValueError, TypeError):
		return None


# --------------------------------------------------------------------------------------
# writers
# --------------------------------------------------------------------------------------


def ensure_document_types(summary, dry_run):
	for row in DOCUMENT_TYPES:
		if frappe.db.exists("ESKD Document Type", row["type_name"]):
			continue
		summary.hit("document types created")
		if dry_run:
			continue
		frappe.get_doc({"doctype": "ESKD Document Type", **row}).insert(ignore_permissions=True)


def upsert_product(name, summary, dry_run, **values):
	name = _norm(name)
	if not name:
		return None
	values = {k: v for k, v in values.items() if v not in (None, "")}
	if frappe.db.exists("ESKD Product", name):
		if values and not dry_run:
			doc = frappe.get_doc("ESKD Product", name)
			doc.update(values)
			doc.save(ignore_permissions=True)
		summary.hit("products updated")
		return name
	summary.hit("products created")
	if dry_run:
		return name
	doc = frappe.get_doc({"doctype": "ESKD Product", "product_name": name, **values})
	doc.insert(ignore_permissions=True)
	return doc.name


def upsert_document(code, summary, dry_run, product=None, **values):
	code = _norm(code)
	if not code:
		return None
	if _is_placeholder(code):
		summary.hit("documents skipped (placeholder code)")
		return None

	values = {k: v for k, v in values.items() if v not in (None, "")}
	values["organization_code"] = values.get("organization_code") or _org_code(code)

	existing = frappe.db.exists("ESKD Document", {"document_code": code, "product": product})
	if existing:
		summary.hit("documents updated")
		if not dry_run:
			doc = frappe.get_doc("ESKD Document", existing)
			doc.update(values)
			doc.save(ignore_permissions=True)
		return existing

	summary.hit("documents created")
	if dry_run:
		return None
	doc = frappe.get_doc(
		{
			"doctype": "ESKD Document",
			"document_code": code,
			"product": product,
			"status": "Active",
			**values,
		}
	)
	doc.insert(ignore_permissions=True)
	return doc.name


def _ensure_parameter(parameter, dry_run):
	if frappe.db.exists("Quality Inspection Parameter", parameter):
		return True
	if dry_run:
		return False
	frappe.get_doc({"doctype": "Quality Inspection Parameter", "parameter": parameter}).insert(
		ignore_permissions=True
	)
	return True


def _free_specification_name(wanted, code):
	"""specification_name is the document ID and must be unique across the catalog."""
	wanted = wanted[:130]
	taken = frappe.db.get_value("Specification", wanted, "specification_code")
	if not taken or taken == code:
		return wanted
	return f"{wanted} ({code})"[:140]


CYRILLIC_ES = "С"
LATIN_ES = "C"


def _same_designation(left, right):
	"""The workbook mixes Cyrillic `С` and Latin `C` as the specification suffix."""
	return (left or "").replace(LATIN_ES, CYRILLIC_ES) == (right or "").replace(LATIN_ES, CYRILLIC_ES)


def upsert_variant(
	code, name, kind, summary, dry_run, attributes=None, parameters=None, components=None, **values
):
	"""Create a catalog entry as a variant so its designation is generated, not typed.

	`code` from the workbook is not written to the document — it is the expectation the
	generated designation is checked against, so a drift between the workbook and the
	template components shows up as a mismatch instead of being silently accepted.
	"""
	from erpnext.manufacturing.eskd_templates import TEMPLATE_BY_KIND

	template = TEMPLATE_BY_KIND.get(kind)
	if not template or not frappe.db.exists("Specification", template):
		return upsert_specification(
			code, name, summary, dry_run, specification_kind=kind, parameters=parameters, **values
		)

	existing = frappe.db.get_value("Specification", {"specification_code": code}, "name")
	if existing:
		summary.hit("variants already present")
		return existing

	summary.hit("variants created")
	if dry_run:
		return None

	doc = frappe.new_doc("Specification")
	doc.variant_of = template
	doc.specification_kind = kind
	doc.specification_name = _free_specification_name(name or code, code)
	doc.update({k: v for k, v in values.items() if v not in (None, "")})
	for attribute, value in attributes or []:
		doc.append("attributes", {"attribute": attribute, "attribute_value": value})
	for role, specification in components or []:
		doc.append("components", {"role": role, "specification": specification})
	for row in _parameter_rows(parameters, dry_run):
		doc.append("parameters", row)
	doc.insert(ignore_permissions=True)

	if not _same_designation(doc.specification_code, code):
		summary.hit("designation mismatches")
		frappe.log_error(
			title="ЄСКД designation mismatch",
			message=f"workbook: {code}\ngenerated: {doc.specification_code}\nvariant: {doc.name}",
		)
	return doc.name


def _parameter_rows(parameters, dry_run):
	rows = []
	for parameter, value, uom in parameters or []:
		if value in (None, ""):
			continue
		if not _ensure_parameter(parameter, dry_run):
			continue
		row = {"parameter": parameter, "value": str(value), "uom": uom or ""}
		numeric = _num(value)
		if numeric is not None:
			row["calculated_value"] = numeric
		rows.append(row)
	return rows


def upsert_specification(code, name, summary, dry_run, parameters=None, **values):
	code = _norm(code)
	name = _norm(name) or code
	if not code:
		return None
	if _is_placeholder(code):
		summary.hit("specifications skipped (placeholder code)")
		return None

	values = {k: v for k, v in values.items() if v not in (None, "")}
	values["organization_code"] = values.get("organization_code") or _org_code(code)

	rows = []
	for parameter, value, uom in parameters or []:
		if value in (None, ""):
			continue
		if not _ensure_parameter(parameter, dry_run):
			continue
		row = {"parameter": parameter, "value": str(value), "uom": uom or ""}
		numeric = _num(value)
		if numeric is not None:
			row["calculated_value"] = numeric
		rows.append(row)

	existing = frappe.db.get_value("Specification", {"specification_code": code}, "name")
	if existing:
		# First writer wins: the authoritative sheets are imported first, so a later
		# draft listing of the same designation only fills in what is still blank and
		# never renames a document that is already linked elsewhere.
		summary.hit("specifications updated")
		if dry_run:
			return existing
		doc = frappe.get_doc("Specification", existing)
		doc.update({k: v for k, v in values.items() if not doc.get(k)})
		if rows and not doc.get("parameters"):
			doc.set("parameters", rows)
		doc.save(ignore_permissions=True)
		return doc.name

	summary.hit("specifications created")
	if dry_run:
		return None
	doc = frappe.get_doc(
		{
			"doctype": "Specification",
			"specification_code": code,
			"specification_name": _free_specification_name(name, code),
			**values,
		}
	)
	if rows:
		doc.set("parameters", rows)
	doc.insert(ignore_permissions=True)
	return doc.name


# --------------------------------------------------------------------------------------
# sheet readers
# --------------------------------------------------------------------------------------


def import_tu_table(wb, summary, dry_run):
	"""`Сводна таблиця ТУ` — one ТУ number per product."""
	ws = wb["Сводная таблиця ТУ"]
	for row in ws.iter_rows(min_row=2, values_only=True):
		product, tu_number, note = _norm(row[1]), _norm(row[2]), _norm(row[3])
		if not product or not tu_number:
			continue
		upsert_product(product, summary, dry_run, tu_number=tu_number, tu_note=note)


def import_register(wb, summary, dry_run):
	"""`Сводная` — the per-product document register, laid out as 3-column blocks."""
	ws = wb["Сводная"]
	grid = [[_norm(c) for c in row] for row in ws.iter_rows(values_only=True)]
	if not grid:
		return
	width = max(len(r) for r in grid)
	for row in grid:
		row.extend([""] * (width - len(row)))

	for block_start in range(0, width, 3):
		product = grid[0][block_start]
		if not product:
			continue
		upsert_product(product, summary, dry_run)
		category = ""
		for row_index, row in enumerate(grid[1:], start=2):
			title, code = row[block_start], row[block_start + 2]
			if not title and not code:
				continue
			if title and not code:
				category = title
				continue
			cell = f"{ws.title}!R{row_index}C{block_start + 3}"
			upsert_document(
				code,
				summary,
				dry_run,
				product=product,
				document_name=title,
				category=category,
				source_sheet=ws.title,
				source_ref=cell,
			)


def import_process_cards(wb, summary, dry_run):
	"""`Технологічні карти` — ТК codes grouped by a product heading row."""
	ws = wb["Технологічні карти"]
	product = None
	for row_index, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
		cells = [_norm(c) for c in row]
		cells.extend([""] * (5 - len(cells)))
		label, code, kind, note = cells[1], cells[2], cells[3], cells[4]
		if label and not code:
			product = upsert_product(label, summary, dry_run)
			continue
		if not code or not product:
			continue
		upsert_document(
			code,
			summary,
			dry_run,
			product=product,
			document_name=note or kind,
			document_type="Технологічна карта",
			category="Технологічні карти",
			source_sheet=ws.title,
			source_ref=f"{ws.title}!R{row_index}C3",
		)


def import_coils(wb, summary, dry_run):
	"""`Специфікація на котушку` — УКРП.200121.002-ХХС catalog."""
	ws = wb["Специфікація на котушку"]
	for row in ws.iter_rows(min_row=5, max_row=64, values_only=True):
		cells = [_norm(c) for c in row]
		cells.extend([""] * (5 - len(cells)))
		ordinal, purpose, code, name = cells[0], cells[1], cells[3], cells[4]
		if not code or not name:
			# reserved-but-unassigned slot
			continue
		upsert_variant(
			code,
			name,
			"Coil",
			summary,
			dry_run,
			attributes=_organisation_attribute(code),
			ordinal=cint(ordinal),
			description=purpose,
			parameters=_coil_parameters(purpose),
		)


ORGANISATION_VALUES = {"УКРП": "Укропчик", "ВРНК": "VARNEX"}


def _organisation_attribute(code):
	from erpnext.manufacturing.eskd_templates import ORGANISATION

	value = ORGANISATION_VALUES.get(_org_code(code))
	return [(ORGANISATION, value)] if value else []


def _coil_parameters(purpose):
	match = COIL_PURPOSE_RE.match(purpose)
	if match:
		spool, fibre, length = match.groups()
		return [
			(PARAM_WINDING_LENGTH, _num(length), "км"),
			(PARAM_SPOOL_DIAMETER, _num(spool), "мм"),
			(PARAM_FIBRE_DIAMETER, _num(fibre), "мм"),
		]
	short = COIL_SHORT_RE.search(purpose)
	if short:
		return [(PARAM_WINDING_LENGTH, _num(short.group(1)), "км")]
	return []


def import_varnex(wb, summary, dry_run):
	"""`ВАРНЕКС` — the ВРНК-branded coil and ground-station lists."""
	ws = wb["ВАРНЕКС"]
	for row in ws.iter_rows(min_row=3, max_row=12, values_only=True):
		cells = [_norm(c) for c in row]
		cells.extend([""] * (6 - len(cells)))
		ordinal, code, name, length, spool, fibre = cells[:6]
		if not code:
			continue
		upsert_variant(
			code,
			name,
			"Coil",
			summary,
			dry_run,
			attributes=_organisation_attribute(code),
			organization_code="ВРНК",
			ordinal=cint(ordinal),
			parameters=[
				(PARAM_WINDING_LENGTH, _num(length), "км"),
				(PARAM_SPOOL_DIAMETER, _num(spool), "мм"),
				(PARAM_FIBRE_DIAMETER, _num(fibre), "мм"),
			],
		)

	for row in ws.iter_rows(min_row=23, max_row=27, values_only=True):
		cells = [_norm(c) for c in row]
		cells.extend([""] * (4 - len(cells)))
		ordinal, code, name, note = cells[:4]
		if not code:
			continue
		upsert_variant(
			code,
			name,
			"Ground Station",
			summary,
			dry_run,
			attributes=_organisation_attribute(code),
			organization_code="ВРНК",
			ordinal=cint(ordinal),
			description=note,
		)


def import_batteries(wb, summary, dry_run):
	"""`Специфікація на батарею` — two side-by-side blocks, УКРП and ВРНК."""
	ws = wb["Специфікація на батарею "]
	blocks = ((0, 1, 4, "УКРП"), (6, 7, 10, "ВРНК"))
	for row in ws.iter_rows(min_row=4, max_row=63, values_only=True):
		cells = [_norm(c) for c in row]
		cells.extend([""] * (11 - len(cells)))
		for ordinal_col, purpose_col, code_col, org in blocks:
			ordinal, purpose, code = cells[ordinal_col], cells[purpose_col], cells[code_col]
			if not code or not purpose:
				# unassigned slot in the reserved range
				continue
			upsert_variant(
				code,
				f"{org} {purpose}",
				"Battery",
				summary,
				dry_run,
				attributes=_organisation_attribute(code),
				organization_code=org,
				ordinal=cint(ordinal),
				description=purpose,
				parameters=[(PARAM_BATTERY_LAYOUT, purpose, "")],
			)


def import_ground_stations(wb, summary, dry_run):
	"""`Специфікація НСУ FO` — УКРП.563562.003-ХХС catalog."""
	ws = wb["Специфікація НСУ FO"]
	for row in ws.iter_rows(min_row=5, max_row=64, values_only=True):
		cells = [_norm(c) for c in row]
		cells.extend([""] * (5 - len(cells)))
		ordinal, purpose, code = cells[0], cells[1], cells[4]
		if not code or not purpose:
			continue
		upsert_variant(
			code,
			purpose,
			"Ground Station",
			summary,
			dry_run,
			attributes=_organisation_attribute(code),
			ordinal=cint(ordinal),
			description=purpose,
		)


# `Сперцифікація на FPV` blocks. The two ПЕРЕЛІК blocks at the bottom of the sheet carry
# the names and parameters that ship with the ТУ, so they are read first and win; the
# working blocks above them only contribute designations the ПЕРЕЛІК lists do not have.
FPV_BLOCKS = (
	{"start": 174, "end": 221, "code": 2, "name": 3, "note": 7, "params": True},
	{"start": 139, "end": 169, "code": 2, "name": 3, "note": 7, "params": True},
	{"start": 84, "end": 134, "code": 2, "name": 4, "note": 3, "params": False},
	{"start": 14, "end": 45, "code": 2, "name": 4, "note": 3, "params": False},
	{"start": 49, "end": 80, "code": 2, "name": None, "note": 3, "params": False},
)


# `УКРП.200121.` + frame(2) + camera(2) + battery ordinal(2) + coil ordinal(4) + `С`
BOARD_CODE_RE = re.compile(
	r"^(?P<org>[А-ЯІЇЄҐA-Z]{4})\.(?P<drone_class>\d{6})\."
	r"(?P<frame>\d{2})(?P<camera>\d{2})(?P<battery>\d{2})(?P<coil>\d{4})[СC]$"
)


def _board_composition(code):
	"""Split a board designation into the attributes and components that generate it."""
	from erpnext.manufacturing.eskd_templates import (
		CAMERA_TYPE,
		DRONE_CLASS,
		FRAME_SIZE,
		ORGANISATION,
		ROLE_BATTERY,
		ROLE_COIL,
	)

	match = BOARD_CODE_RE.match(_norm(code))
	if not match:
		return None

	organisation = ORGANISATION_VALUES.get(match.group("org"))
	drone_class = _attribute_value_for_abbr(DRONE_CLASS, match.group("drone_class"))
	frame = _attribute_value_for_abbr(FRAME_SIZE, match.group("frame"))
	camera = _attribute_value_for_abbr(CAMERA_TYPE, match.group("camera"))
	if not all((organisation, drone_class, frame, camera)):
		return None

	components = []
	for role, kind, ordinal in (
		(ROLE_BATTERY, "Battery", match.group("battery")),
		(ROLE_COIL, "Coil", match.group("coil")),
	):
		entry = _catalog_entry(kind, match.group("org"), ordinal)
		if entry:
			components.append((role, entry))

	return {
		"attributes": [
			(ORGANISATION, organisation),
			(DRONE_CLASS, drone_class),
			(FRAME_SIZE, frame),
			(CAMERA_TYPE, camera),
		],
		"components": components,
	}


def _attribute_value_for_abbr(attribute, abbr):
	return frappe.db.get_value("Item Attribute Value", {"parent": attribute, "abbr": abbr}, "attribute_value")


def _catalog_entry(kind, organization_code, ordinal):
	return frappe.db.get_value(
		"Specification",
		{
			"specification_kind": kind,
			"organization_code": organization_code,
			"ordinal": cint(ordinal),
		},
		"name",
	)


def import_boards(wb, summary, dry_run):
	"""`Сперцифікація на FPV` — the per-airframe board specification lists."""
	ws = wb["Сперцифікація на FPV"]
	grid = [[_norm(c) for c in row] for row in ws.iter_rows(values_only=True)]
	width = max(len(r) for r in grid)
	for row in grid:
		row.extend([""] * (width - len(row)))

	for block in FPV_BLOCKS:
		for row in grid[block["start"] - 1 : block["end"]]:
			code = row[block["code"]]
			if not code:
				continue
			note = row[block["note"]] if block["note"] is not None else ""
			name = row[block["name"]] if block["name"] is not None else ""
			parameters = []
			if block["params"]:
				parameters = [
					(PARAM_WINDING_LENGTH, _num(row[4]), "км"),
					(PARAM_CAMERA_CHANNEL, row[5], ""),
					(PARAM_CAMERA_SIGNAL, row[6], ""),
				]
			composition = _board_composition(code)
			if composition is None:
				# Radio boards such as `УКРП.463145.106C/15` do not follow the assembled
				# grammar — keep them as plain catalog entries.
				upsert_specification(
					code,
					name or note,
					summary,
					dry_run,
					specification_kind="Board",
					description=note,
					parameters=parameters,
				)
				continue
			upsert_variant(
				code,
				name or note,
				"Board",
				summary,
				dry_run,
				attributes=composition["attributes"],
				components=composition["components"],
				description=note,
				parameters=parameters,
			)


def import_modifications(wb, summary, dry_run):
	"""`Відомість модифікацій ТУ14` — the numbered modification list of a БпАК.

	The row axis is text (modification number -> board specification); the intersections
	are **cell fills**, not values — a solid-filled cell under a ground-station column is
	the pairing. Reading only values silently loses all of them.
	"""
	ws = wb["Відомість модифікацій ТУ14"]
	grid = [[_norm(c) for c in row] for row in ws.iter_rows(values_only=True)]
	if len(grid) < 5:
		return

	product = _modification_product(grid[2][0])
	if not product:
		summary.hit("modification sheets skipped (no product)")
		return
	upsert_product(product, summary, dry_run)

	header_row = 4
	columns = {
		col: grid[header_row - 1][col - 1]
		for col in range(1, len(grid[header_row - 1]) + 1)
		if grid[header_row - 1][col - 1]
	}

	for row_index in range(header_row + 1, len(grid) + 1):
		row = grid[row_index - 1]
		number = _modification_number(row[0])
		board_code = row[2] if len(row) > 2 else ""
		if not number or not board_code:
			continue
		upsert_combination(
			product,
			number,
			board_code,
			row[1],
			_marked_ground_station(ws, row_index, columns),
			summary,
			dry_run,
		)


def _marked_ground_station(ws, row_index, columns):
	"""Return the ground-station designation whose cell is filled on this row."""
	for col, code in columns.items():
		if ws.cell(row_index, col).fill.patternType == "solid":
			return code
	return ""


def _modification_product(title):
	"""`Відомість модифікацій БпАК Укропчик 15 FO УКРП.463145.006ВМ` -> `Укропчик 15 FO`."""
	text = _norm(title)
	marker = "БпАК "
	if marker not in text:
		return ""
	tail = text.split(marker, 1)[1]
	return re.sub(r"\s+[А-ЯІЇЄҐA-Z]{4}\.\S+$", "", tail).strip()


def _modification_number(label):
	match = re.search(r"(\d+)", _norm(label))
	return cint(match.group(1)) if match else 0


def upsert_combination(product, number, board_code, board_name, gs_code, summary, dry_run):
	board = _specification_by_code(board_code)
	if not board:
		summary.hit("combinations skipped (board specification not in catalog)")
		return None

	ground_station = _specification_by_code(gs_code) if gs_code else None
	if gs_code and not ground_station:
		summary.hit("intersections whose ground station is not in the catalog")
	elif ground_station:
		summary.hit("intersections read from cell fills")

	if not ground_station:
		# without a marked intersection there is no pairing to record
		return None

	existing = frappe.db.exists(
		"Specification", {"specification_kind": "BpAK", "product": product, "ordinal": number}
	)
	if existing:
		summary.hit("combinations already present")
		return existing

	summary.hit("combinations created")
	if dry_run:
		return None

	from erpnext.manufacturing.eskd_templates import ROLE_BOARD, ROLE_GROUND_STATION

	doc = frappe.new_doc("Specification")
	doc.variant_of = "Специфікація БпАК"
	doc.specification_kind = "BpAK"
	doc.product = product
	doc.ordinal = number
	doc.specification_name = f"{product} — модифікація {number}"
	doc.specification_code = f"{board_code} / {gs_code}"
	doc.description = board_name
	doc.append("components", {"role": ROLE_BOARD, "specification": board})
	doc.append("components", {"role": ROLE_GROUND_STATION, "specification": ground_station})
	doc.insert(ignore_permissions=True)
	return doc.name


def _specification_by_code(code):
	"""Look up a designation, tolerating the workbook's mixed Cyrillic `С` / Latin `C`."""
	code = _norm(code)
	if not code:
		return None
	found = frappe.db.get_value("Specification", {"specification_code": code}, "name")
	if found:
		return found
	swapped = code.replace(LATIN_ES, CYRILLIC_ES)
	if swapped != code:
		found = frappe.db.get_value("Specification", {"specification_code": swapped}, "name")
	return found


SHEET_IMPORTERS = {
	"tu": import_tu_table,
	"register": import_register,
	"process_cards": import_process_cards,
	"coils": import_coils,
	"varnex": import_varnex,
	"batteries": import_batteries,
	"ground_stations": import_ground_stations,
	"boards": import_boards,
	"modifications": import_modifications,
}


def run(path, dry_run=True, only=None):
	"""Import the ЄСКД workbook. Pass dry_run=False to actually write."""
	import openpyxl

	dry_run = bool(dry_run)
	wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
	summary = Summary()

	ensure_document_types(summary, dry_run)
	if not dry_run:
		from erpnext.manufacturing.eskd_templates import setup

		setup()

	names = [only] if isinstance(only, str) else (only or list(SHEET_IMPORTERS))
	for key in names:
		importer = SHEET_IMPORTERS.get(key)
		if not importer:
			frappe.throw(f"Unknown ЄСКД importer: {key}")
		importer(wb, summary, dry_run)

	if not dry_run:
		frappe.db.commit()

	print(("DRY RUN — nothing written" if dry_run else "IMPORT COMMITTED") + f" ({path})")
	summary.show()
	return summary.as_dict()
